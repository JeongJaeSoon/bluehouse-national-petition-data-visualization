from openpyxl import load_workbook
from konlpy.tag import Kkma
import collections

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.font_manager as font_manage

# 엑셀 파일 불러와서 열기
read_workbook = load_workbook("./bluehouse.xlsx")
read_worksheet = read_workbook.active

list_excel = []

for i in range(1, 151):
    list_excel.append(read_worksheet.cell(i, 1).value.strip())

kkma = Kkma()
list_temp = []

# 단어를 구분하여 저장
for row in list_excel:
    list_temp = list_temp + kkma.nouns(row)

list_result = []

# 1음절 단어는 제외하고 저장 후, top 20 단어 저장
for check in list_temp:
    if len(check) > 1:
        list_result.append(check)

list_data = collections.Counter(list_result).most_common(20)

list_string = []
list_number = []

# 차트에 사용할 데이터 set
for string, number in list_data:
    list_string.append(string)
    list_number.append(number)

label = list_string
index = np.arange(len(label))

# 한글 폰트 설정
font_name = font_manage.FontProperties(fname="./D2Coding.ttc").get_name()
plt.rc('font', family=font_name)
plt.bar(index, list_number)

plt.title("my title")
plt.xlabel("x-label")
plt.ylabel("y-label")
plt.xticks(index, label)
plt.xticks(rotation=60)

plt.show()