from openpyxl import load_workbook
from konlpy.tag import Kkma
import collections

read_workbook = load_workbook("./bluehouse.xlsx")
read_worksheet = read_workbook.active

excel_list = []

for i in range(1, 151):
    excel_list.append(read_worksheet.cell(i, 1).value.strip())

kkma = Kkma()
temp_list = []

for row in excel_list:
    temp_list = temp_list + kkma.nouns(row)

result_list = []

for check in temp_list:
    if len(check) > 1:
        result_list.append(check)

print(collections.Counter(result_list).most_common(10))
