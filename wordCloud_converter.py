from openpyxl import load_workbook
from konlpy.tag import Kkma

import matplotlib.pyplot as plt
from wordcloud import WordCloud

from os import path
from PIL import Image
import numpy as np
import os

read_workbook = load_workbook("./bluehouse.xlsx")
read_worksheet = read_workbook.active

list_excel = []

for i in range(1, 151):
    list_excel.append(read_worksheet.cell(i, 1).value.strip())

kkma = Kkma()
list_temp = []

for row in list_excel:
    list_temp = list_temp + kkma.nouns(row)

list_result = []

for check in list_temp:
    if len(check) > 1:
        list_result.append(check)

last_text = ""

for result in list_result:
    last_text = f"{last_text} {result}"

# 이미지 불러오기
d = path.dirname(__file__) if "__file__" in locals() else os.getcwd()
mask = np.array(Image.open(path.join(d, "./cloud.png")))

# Word Cloud 생성
wordcloud = WordCloud(
    font_path="./D2Coding.ttc",
    background_color="white",
    mask=mask,
).generate(last_text)
plt.figure(figsize=(12, 12))
plt.imshow(wordcloud, interpolation="bilinear")
plt.axis("off")
plt.show()
